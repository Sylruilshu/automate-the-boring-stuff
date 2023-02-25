import openpyxl


source_workbook = openpyxl.load_workbook("cell_inverter.xlsx")
source_sheet = source_workbook["Sheet"]

target_workbook = openpyxl.Workbook()
target_sheet = target_workbook["Sheet"]

source_sheet_data = []

for x in range(source_sheet.max_column):
    column_data = []
    for y in range(source_sheet.max_row):
        source_cell_data = source_sheet.cell(row=y + 1, column=x + 1)
        column_data.append(source_cell_data)

    source_sheet_data.append(column_data)

for y in range(source_sheet.max_column):
    for x in range(source_sheet.max_row):
        target_sheet.cell(row=y + 1, column=x + 1).value = source_sheet_data[y][x].value


target_workbook.save("cell_inverter_inverted.xlsx")
