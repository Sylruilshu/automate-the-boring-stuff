import openpyxl, click
from openpyxl.utils import get_column_letter


EXCEL_FILE_EXTENTION = ".xlsx"
TARGET_FILE_PREFIX = "_new"


def blank_row_inserter(
    filename: str, start_index: int, blank_rows_to_insert: int
) -> None:
    source_workbook = openpyxl.load_workbook(filename + EXCEL_FILE_EXTENTION)
    source_sheet = source_workbook["Sheet"]

    target_workbook = openpyxl.Workbook()
    target_sheet = target_workbook["Sheet"]

    max_column_letter = get_column_letter(source_sheet.max_column)

    for row in source_sheet["A1":f"{max_column_letter}{start_index}"]:
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

    for row in source_sheet[
        f"A{start_index + 1}":f"{max_column_letter}{source_sheet.max_row}"
    ]:
        for cell in row:
            target_row = cell.row + blank_rows_to_insert
            # new_position = (
            #     f"{cell.column_letter}"
            #     + f"{int(cell.coordinate[1:]) + blank_rows_to_insert}"
            # )
            # target_sheet[new_position].value = cell.value
            target_sheet.cell(row=target_row, column=cell.column).value = cell.value

    target_workbook.save(filename + TARGET_FILE_PREFIX + EXCEL_FILE_EXTENTION)


@click.command()
@click.option("--filename", "-f", type=str, required=True, help="The name of the file")
@click.option(
    "--row",
    "-r",
    type=int,
    default=1,
    help="The row before blank rows will be inserted",
)
@click.option(
    "--blank_rows_to_insert",
    "-b",
    type=int,
    default=1,
    help="The amount of rows to insert",
)
def main(filename: str, row: int, blank_rows_to_insert: int) -> None:
    """
    A utility to insert blank rows into an excel spreadsheet
    """
    blank_row_inserter(filename, row, blank_rows_to_insert)


main()
