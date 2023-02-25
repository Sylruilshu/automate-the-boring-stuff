import openpyxl, sys
from openpyxl.styles import Font


multiplication_table_number = int(sys.argv[1]) + 1

workbook = openpyxl.Workbook()

sheet = workbook["Sheet"]
sheet.title = "Multiplication-Table"

font_bold = Font(bold=True)

for y in range(1, multiplication_table_number):
    for x in range(1, multiplication_table_number):
        # Row 1 (top row)
        sheet.cell(row=1, column=x + 1).value = x
        sheet.cell(row=1, column=x + 1).font = font_bold
        # Column A (first column)
        sheet.cell(row=x + 1, column=1).value = x
        sheet.cell(row=x + 1, column=1).font = font_bold

        sheet.cell(row=y + 1, column=x + 1).value = f"=(A{x+1}*{y})"

workbook.save("multiplication_table.xlsx")
