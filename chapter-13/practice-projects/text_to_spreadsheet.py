import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter


workbook = openpyxl.Workbook()
sheet = workbook.active

files = []

for path in Path("text-to-spreadsheet").glob("*.txt"):
    files.append(path.name)

for column, file in enumerate(files, 1):
    file_data = open(f"text-to-spreadsheet/{file}", "r")
    lines = file_data.readlines()

    column_width = len(max(lines, key=len))
    sheet.column_dimensions[get_column_letter(column)].width = column_width

    file_data.close()

    for row, line in enumerate(lines, 1):
        sheet.cell(row, column).value = line.strip("\n")

workbook.save("text_to_spreadsheet.xlsx")
